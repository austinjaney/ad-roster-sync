#!/usr/bin/env python3
"""Build AD reconciliation artifacts from roster + AzureAD + Kandji + passwordless exports.

Matching model reflects NCC email convention:
  primary   : <first-initial><lastname>        e.g. jsmith
  exception : <first>.<lastname>               used when flast collided
  exception : <firstname>                      legacy single-name addresses
Every candidate is cross-checked against the AD displayName so an initial
collision (one initial matching two different people) is flagged, not applied.
"""
import csv, re, unicodedata, collections, difflib, os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

SRC = OUT = os.environ.get("ROSTER_DIR", os.path.dirname(os.path.abspath(__file__)))

def fold(s): return unicodedata.normalize("NFKD", str(s or "")).encode("ascii", "ignore").decode()
def toks(s): return re.sub(r"[^a-z ]", " ", fold(s).lower()).split()
def alpha(s): return re.sub(r"[^a-z]", "", fold(s).lower())
def nkey(s): return " ".join(toks(s))

NICK = {
    "christopher": {"chris"}, "jennifer": {"jenn", "jen"}, "samuel": {"sam"},
    "joshua": {"josh"}, "vincent": {"vince"}, "phillip": {"philip", "phil"},
    "kirkland": {"kirk"}, "patricia": {"patti", "pat", "patty"}, "stanley": {"stan"},
    "david": {"dave"}, "zachary": {"zach", "zack"}, "timothy": {"tim"},
    "benjamin": {"ben"}, "jonathan": {"jon"}, "michael": {"mike"},
    "richard": {"rich", "rick"}, "allison": {"alli", "allie", "ally"},
    "nathan": {"nate"}, "elliott": {"elliot"}, "deborah": {"debbie", "deb"},
    "thomas": {"tom", "tommy"}, "alexander": {"alex"}, "matthew": {"matt"},
    "robert": {"rob", "robby", "bob"}, "sara": {"sarah"},
    "gregory": {"greg"}, "bradley": {"brad", "bj"},
    
    "daniel": {"dan", "danny", "dj"}, "maleine": {"melanie"},
    "william": {"will", "bill"}, "andrew": {"andy", "drew"}, "jacob": {"jake"},
    "joseph": {"joe", "joey"}, "kathryn": {"kate", "katie", "kathy"},
    "rebecca": {"becky", "becca"}, "suzanne": {"sue"}, "gerardo": {"junior"},
    "melanie": {"mel"}, "jeffrey": {"jeff"}, "steven": {"steve"}, "stephen": {"steve"},
    "kenneth": {"ken"}, "ronald": {"ron"}, "donald": {"don"}, "charles": {"charlie", "chuck"},
    "edward": {"ed", "eddie"}, "anthony": {"tony"}, "nicholas": {"nick"},
}
def first_agree(a, b, alias=""):
    """0..100 agreement between two given names."""
    a, b = alpha(a), alpha(b)
    if not a or not b: return 0
    if a == b: return 100
    for x, y in ((a, b), (b, a)):
        if y in NICK.get(x, set()): return 92
    if alias:
        al = alpha(alias)
        if al == b or b in NICK.get(al, set()) or al in NICK.get(b, set()): return 88
        if al and b and al[0] == b[0] and difflib.SequenceMatcher(None, al, b).ratio() > .7: return 80
    if a.startswith(b) or b.startswith(a): return 85          # Tom/Thomas, Alli/Allison
    r = difflib.SequenceMatcher(None, a, b).ratio()
    if r > .82: return 78                                      # Keegan/Keagan, Elliott/Elliot
    if r > .70: return 60
    return 20 if a[0] == b[0] else 0

def last_agree(a, b):
    a, b = alpha(a), alpha(b)
    if not a or not b: return 0
    if a == b: return 100
    if a.startswith(b) or b.startswith(a): return 85
    r = difflib.SequenceMatcher(None, a, b).ratio()
    return 75 if r > .85 else (50 if r > .7 else 0)

def parse_roster_name(raw):
    s = " ".join(str(raw).split()); alias = ""
    m = re.search(r"\(([^)]*)\)", s)
    if m:
        alias = m.group(1).strip(); s = re.sub(r"\([^)]*\)", " ", s)
    s = " ".join(s.split())
    if "," in s:
        last, first = [p.strip() for p in s.split(",", 1)]
    else:
        p = s.split(); last, first = (p[0], " ".join(p[1:])) if len(p) > 1 else (s, "")
    return first.strip(), last.strip(), alias

# ---------------- load ----------------
wb = openpyxl.load_workbook(os.path.join(SRC, os.environ.get("ROSTER_FILE", "Staff-Roster.xlsx")))
ws = wb["Sheet1"]
roster = []
for i, r in enumerate(ws.iter_rows(min_row=2, max_col=6, values_only=True), start=2):
    if not r[0] or not str(r[0]).strip(): continue
    roster.append({"row": i, "raw": " ".join(str(r[0]).split()), "campus": (r[1] or "").strip(),
                   "reports_to": (r[2] or "").strip(), "dept": (r[3] or "").strip(),
                   "title": (r[4] or "").strip(), "ftpt": (r[5] or "").strip()})

aad_all = list(csv.DictReader(open(os.path.join(SRC, "AzureAD-exportUsers.csv"), encoding="utf-8-sig")))
pool = [a for a in aad_all
        if a["userPrincipalName"].lower().endswith("@example.org")
        and "#ext#" not in a["userPrincipalName"].lower() and a["userType"] == "Member"]
for a in pool:
    a["_local"] = a["userPrincipalName"].split("@")[0].lower()
    a["_toks"] = toks(a["displayName"])

by_sam_pin = {}
by_local = collections.defaultdict(list)
by_last = collections.defaultdict(list)
by_full = collections.defaultdict(list)
for a in pool:
    by_local[a["_local"]].append(a)
    by_sam_pin[a["_local"]] = a
    if a["_toks"]:
        by_last[alpha(a["_toks"][-1])].append(a)
        by_full[" ".join(a["_toks"])].append(a)

devices = list(csv.DictReader(open(os.path.join(SRC, "All-Devices.csv"), encoding="utf-8-sig")))
pwless = list(csv.DictReader(open(os.path.join(SRC, "Passwordless-Pending.csv"), encoding="utf-8-sig")))
pw_upn = {p["userPrincipalName"].strip().lower() for p in pwless if p["userPrincipalName"].strip()}

# ---------------------------------------------------------------------------
# SITE CONFIGURATION - replace all four maps with your own.
#
# CAMPUS_COMPANY and CAMPUS_OU can legitimately differ: a site with no OU of its own may be
# filed with central staff without being the central site, and writing that into the address
# book would be wrong.
# ---------------------------------------------------------------------------
CAMPUS_OU      = {"HQ": "HQ", "North": "North", "South": "South", "Central": "Central"}
CAMPUS_COMPANY = {"HQ": "HQ Campus", "North": "North Campus", "South": "South Campus",
                  "Central": "Central Team"}
CAMPUS_DISPLAY = {"HQ": "HQ", "North": "North", "South": "South", "Central": "Central"}

REMAPPED = {
    # Roster sites with no OU of their own. Reported, not silently remapped.
}

# ---------------------------------------------------------------------------
# HUMAN DECISIONS. Shipped empty; fill in as you review.
#
# These exist so a decision made once survives a rebuild. Record WHY next to it - otherwise
# the matcher re-offers the same wrong near-miss every cycle and the reviewer re-derives the
# same conclusion. Key is the roster row number.
# ---------------------------------------------------------------------------

# Identities scoring cannot resolve. The case this exists for: two people sharing a display
# name AND a company - a relative pair, say. No rule separates them.
PINS = {
    #   46: ("jsmith2", "the North site lead; 'jsmith' is a different J. Smith"),
}

# Rows settled as having no account, so the matcher stops offering the same near-miss -
# typically a same-surname colleague who is a different person.
NO_ACCOUNT_ROWS = {
    #   273: "no login; the near-miss is a different person with the same surname",
}

# Roster names replaced wholesale after confirmation - a marriage, a legal name change, or a
# deliberate choice to follow AD. Distinct from the automatic spelling pass.
ROSTER_NAME_OVERRIDES = {
    #   14: ("Newname, Jane", "married; AD surname is current and the roster was stale"),
}

# People holding more than one role. AD stores one title, so the rest are suppressed. Forced
# rather than inferred, so tie-breaking never quietly reassigns which role AD shows.
SECONDARY_ROLE_ROWS = {
    #   216: "second role; row 208 is the primary",
}

# ---------------- match ----------------
def score(p, a):
    """Composite 0..100 for roster person p vs AD account a."""
    if not a["_toks"]: return 0, "no displayName"
    ad_first, ad_last = a["_toks"][0], a["_toks"][-1]
    # multi-word given names on either side: compare full leading run
    r_first_full = alpha(p["first"]).replace(" ", "")
    ad_first_full = alpha(" ".join(a["_toks"][:-1]))
    fs = max(first_agree(p["first"], ad_first, p["alias"]),
             first_agree(r_first_full, ad_first_full, p["alias"]))
    ls = last_agree(p["last"], ad_last)
    if ls == 0: return 0, f"surname mismatch ({ad_last})"
    if fs == 0: return 0, f"given-name mismatch ({ad_first})"
    return round(0.55 * fs + 0.45 * ls), ""

seen = {}
for p in roster:
    p["first"], p["last"], p["alias"] = parse_roster_name(p["raw"])
    p["ou"] = CAMPUS_OU.get(p["campus"], "")
    p["ou_note"] = f"{p['campus']} -> {p['ou']} (no dedicated OU; remapped)" if p["campus"] in REMAPPED else ""
    k = nkey(f"{p['first']} {p['last']}")
    p["dupe_of"] = seen.get(k, ""); seen.setdefault(k, p["row"])
    p["pin_note"] = ""

    fa, la = alpha(p["first"]), alpha(p["last"])
    ai = alpha(p["alias"])
    # A hyphenated or two-part surname on the roster is often carried into AD as just one
    # of its parts - 'French-Apana' becomes Apana, 'Soberon-West' becomes West.
    surnames = [la] + [alpha(x) for x in re.split(r"[-\s]+", p["last"]) if alpha(x) and alpha(x) != la]
    givens = [fa] + ([ai] if ai else [])
    pats = []
    for g in givens:
        if not g: continue
        tag = "alias" if g == ai and g != fa else ""
        for sn in surnames:
            if not sn: continue
            part = "-part" if sn != la else ""
            pats += [(g[0] + sn, f"{tag}flast{part}"), (g + "." + sn, f"{tag}first.last{part}"),
                     (g + sn[0], f"{tag}first+lastinitial{part}")]
        pats.append((g, f"{tag}firstonly"))
    pats = [(loc, kind.lstrip()) for loc, kind in pats]
    cands = []
    for loc, kind in pats:
        for a in by_local.get(loc, []):
            sc, why = score(p, a)
            cands.append({"a": a, "score": sc, "via": f"UPN:{kind}", "why": why})
    for a in by_full.get(nkey(f"{p['first']} {p['last']}"), []):   # display-name safety net
        if not any(c["a"] is a for c in cands):
            sc, why = score(p, a)
            cands.append({"a": a, "score": sc, "via": "displayName", "why": why})
    for sn in surnames:                                            # surname sweep for renamed UPNs
        for a in by_last.get(sn, []):
            if not any(c["a"] is a for c in cands):
                sc, why = score(p, a)
                if sc >= 70: cands.append({"a": a, "score": sc, "via": "surname-sweep", "why": why})

    # Nothing yet? The surname may simply be spelled differently on the roster than in AD
    # (Smyth/Smythe, Anderson/Andersen). Compare UPN local parts and require the
    # given name to agree, then cap the score so a spelling variant always needs a human.
    if not [c for c in cands if c["score"] > 0]:
        # Every naming pattern gets a near-miss pass, not just flast. An account may hold a
        # first.last address while the roster misspells the surname, so a flast-only fallback
        # never reaches it.
        targets = []
        for g in givens:
            if not g: continue
            for sn in surnames:
                if not sn: continue
                targets += [g[0] + sn, g + "." + sn, g + sn[0]]
        for target in dict.fromkeys(targets):
            for a in pool:
                lp = a["_local"]
                if not a["_toks"] or lp[:1] != target[:1]: continue
                r = difflib.SequenceMatcher(None, lp, target).ratio()
                if r < 0.86: continue
                if first_agree(alpha(p["first"]), a["_toks"][0], p["alias"]) < 70: continue
                if any(c["a"] is a for c in cands): continue
                cands.append({"a": a, "score": min(85, round(100 * r)),
                              "via": f"spelling:{lp}~{target}",
                              "why": f"AD spells the surname '{a['_toks'][-1]}', "
                                     f"roster spells it '{p['last']}'"})

    cands.sort(key=lambda c: (-c["score"], c["a"]["_local"]))
    p["cands"] = cands
    top = cands[0] if cands else None
    if not cands or top["score"] == 0:
        p["tier"], p["conf"], p["match"] = "NO_STAFF_ACCOUNT", 0, None
    elif top["score"] >= 90 and (len(cands) == 1 or cands[1]["score"] < 90):
        p["tier"], p["conf"], p["match"] = "AUTO", top["score"], top
    elif top["score"] >= 90:
        p["tier"], p["conf"], p["match"] = "AMBIGUOUS", top["score"], top
    else:
        p["tier"], p["conf"], p["match"] = "REVIEW", top["score"], top

# reject an AUTO if two roster people claim the same account
claim = collections.defaultdict(list)
for p in roster:
    if p["tier"] == "AUTO": claim[p["match"]["a"]["_local"]].append(p)
for loc, ps in claim.items():
    if len(ps) > 1:
        for p in ps:
            if nkey(f"{p['first']} {p['last']}") != nkey(p["match"]["a"]["displayName"]):
                p["tier"], p["conf"] = "AMBIGUOUS", min(p["conf"], 89)

# ---------------- apply human decisions ----------------
for p in roster:
    if p["row"] in NO_ACCOUNT_ROWS:
        p["match"], p["tier"], p["conf"] = None, "NO_STAFF_ACCOUNT", 0
        p["pin_note"] = "settled by review: " + NO_ACCOUNT_ROWS[p["row"]]
        continue
    if p["row"] not in PINS: continue
    sam, why = PINS[p["row"]]
    a = by_sam_pin.get(sam)
    if not a:
        p["pin_note"] = f"PIN FAILED - '{sam}' not in the Entra export"
        continue
    p["match"] = {"a": a, "score": 100, "via": "pinned", "why": ""}
    p["tier"], p["conf"], p["pin_note"] = "AUTO", 100, why

# ---------------- dual-role resolution ----------------
# 11 people hold two or three roles on the roster. AD carries one title/department per
# object, so pick a primary: full-time beats part-time, then the earliest roster row.
claimed = collections.defaultdict(list)
for p in roster:
    p["role_rank"] = 0
    if p["tier"] == "AUTO": claimed[p["match"]["a"]["_local"]].append(p)
dual = []
for loc, ps in claimed.items():
    if len(ps) < 2: continue
    # A row named in SECONDARY_ROLE_ROWS always sorts last, whatever the FT/PT tie-break
    # would have produced. Everything else: full-time first, then earliest roster row.
    ps.sort(key=lambda q: (1 if q["row"] in SECONDARY_ROLE_ROWS else 0,
                           0 if q["ftpt"].upper() == "FT" else 1, q["row"]))
    for i, q in enumerate(ps): q["role_rank"] = i
    for q in ps:
        if q["row"] in SECONDARY_ROLE_ROWS and q["role_rank"] == 0:
            raise SystemExit(f"row {q['row']} is marked secondary but sorted primary - check SECONDARY_ROLE_ROWS")
    dual.append((loc, ps))

# ---------------- manager resolution ----------------
# The 'Reports To' column is typed free-hand, so it carries spellings the Name column
# does not. Corrections go here rather than into a fuzzy rule, so a near-miss can never
# silently point a whole department at the wrong manager.
MANAGER_ALIASES = {
    # The "Reports To" column is typed free-hand, so it carries spellings the Name column
    # does not. Corrections go here rather than into a fuzzy rule, so a near-miss can never
    # silently point a whole department at the wrong manager.
    #   "jane smyth": "jane smythe",
}
NON_PERSON_MANAGERS = {"board"}     # governance bodies, not AD objects     # governance body, not an AD object

by_sam = {a["_local"]: a for a in pool}
def resolve_mgr(name):
    t = toks(name)
    if not t: return "", ""
    key = " ".join(t)
    if key in NON_PERSON_MANAGERS:
        return "", "manager is not a person"
    if key in MANAGER_ALIASES:
        t = toks(MANAGER_ALIASES[key]); key = " ".join(t)
    hit = by_full.get(key)
    if hit:
        staff = [a for a in hit if a["_local"] in MATCHED_SAMS]
        if len(hit) > 1 and len(staff) == 1:
            return staff[0]["_local"], (f"'{name}' matches {len(hit)} accounts; picked the one on "
                                        f"the roster ({staff[0]['_local']})")
        if len(hit) > 1 and len(staff) != 1:
            return hit[0]["_local"], (f"'{name}' matches {len(hit)} accounts and the roster does not "
                                      f"disambiguate - verify manually")
        return hit[0]["_local"], ""
    if len(t) >= 2:
        for loc in (alpha(t[0])[:1] + alpha(t[-1]), alpha(t[0]) + "." + alpha(t[-1]), alpha(t[0])):
            if loc in by_sam: return by_sam[loc]["_local"], ""
    # roster-internal fallback: manager may themselves be a matched roster person
    for q in roster:
        if nkey(f"{q['first']} {q['last']}") == " ".join(t) and q.get("match") and q["tier"] == "AUTO":
            return q["match"]["a"]["_local"], ""
    # last resort: unique surname near-miss with the same first name. Reported, not silent.
    if len(t) >= 2:
        near = [a for a in pool if a["_toks"]
                and alpha(a["_toks"][0]) == alpha(t[0])
                and difflib.SequenceMatcher(None, alpha(a["_toks"][-1]), alpha(t[-1])).ratio() > .88]
        if len(near) == 1:
            return near[0]["_local"], (f"manager matched on a near-miss surname: "
                                       f"'{name}' -> '{near[0]['displayName']}'")
    return "", "manager not found in AD"

# Accounts already confirmed as roster staff. When a manager's display name is ambiguous,
# the staff account wins. Where a display name resolves to two accounts - a relative pair,
# say - the one on the roster is the one who manages people. Otherwise the reports get
# whichever account the export happened to list first.
MATCHED_SAMS = {p["match"]["a"]["_local"] for p in roster if p["tier"] == "AUTO"}

mgr_cache = {}
for p in roster:
    if p["reports_to"] not in mgr_cache:
        mgr_cache[p["reports_to"]] = resolve_mgr(p["reports_to"])
    p["mgr_sam"], p["mgr_note"] = mgr_cache[p["reports_to"]]

# ---------------- sign-in recency + ambiguous display names ----------------
import datetime, json
EXPORT_DATE = datetime.date.today()          # set explicitly if the export is older
STALE_DAYS = 180

# Interactive sign-ins only. lastNonInteractiveSignInDateTime records token refreshes - a
# phone still collecting mail, a background service - not a person at a keyboard, and it
# keeps departed accounts looking alive: an account with no interactive sign-in for two
# years can still show non-interactive traffic from last month, which reads as active.
INTERACTIVE_FIELDS = ("lastSuccessfulSignInDateTime", "lastSignInDateTime")
NONINTERACTIVE_FIELD = "lastNonInteractiveSignInDateTime"

def _sidate(s, k):
    v = s.get(k)
    return datetime.date.fromisoformat(v[:10]) if v else None

def signin_dates(a):
    """-> (last interactive, last non-interactive)"""
    try: s = json.loads(a["signInActivity"]) if a["signInActivity"] else {}
    except Exception: return None, None
    inter = [d for d in (_sidate(s, k) for k in INTERACTIVE_FIELDS) if d]
    return (max(inter) if inter else None), _sidate(s, NONINTERACTIVE_FIELD)

for a in pool:
    a["_last"], a["_last_noninteractive"] = signin_dates(a)
    a["_age"] = (EXPORT_DATE - a["_last"]).days if a["_last"] else None
    a["_noninteractive_only"] = (
        (a["_age"] is None or a["_age"] >= STALE_DAYS)
        and a["_last_noninteractive"] is not None
        and (EXPORT_DATE - a["_last_noninteractive"]).days < STALE_DAYS)

# ---- disabled accounts, supplied separately ----
# The Entra export carries no accountEnabled column - onPremisesSyncEnabled describes
# directory sync, not account state - so disabled status cannot be derived from it. Generate
# the file on a domain-joined workstation with Export-ADDisabledAccounts.ps1 and drop it
# beside this script; without it every account is treated as enabled and a note says so.
DISABLED_FILE = os.path.join(SRC, "AD-Disabled-Accounts.csv")
disabled_sams = set()
disabled_detail = {}
disabled_known = os.path.exists(DISABLED_FILE)
if disabled_known:
    for d in csv.DictReader(open(DISABLED_FILE, encoding="utf-8-sig")):
        keys = set()
        for key in ("SamAccountName", "UserPrincipalName"):
            v = (d.get(key) or "").strip().lower()
            if v: keys.add(v.split("@")[0])
        # Carry the AD-side context through. Descriptions on these rows often hold the
        # offboarding ticket and date, which is exactly what a reviewer needs.
        ou = (d.get("CurrentOU") or "").split(",")[0].replace("OU=", "").replace("CN=", "")
        det = []
        if d.get("Title"): det.append("AD title: " + d["Title"].strip())
        if d.get("Description"): det.append("AD note: " + d["Description"].strip())
        if ou: det.append("in " + ou)
        if d.get("LastLogonDate"): det.append("last logon " + d["LastLogonDate"])
        for k in keys:
            disabled_sams.add(k)
            disabled_detail[k] = " | ".join(det)
print(("disabled list loaded: %d accounts" % len(disabled_sams)) if disabled_known
      else "NOTE: AD-Disabled-Accounts.csv absent - no account can be marked disabled")

# display names carried by more than one internal account cannot be attributed
# from Kandji's free-text "Device User" field alone
dispcount = collections.Counter(nkey(a["displayName"]) for a in pool)
AMBIG_DISPLAY = {k for k, v in dispcount.items() if v > 1}

# ---------------- assets ----------------
# Only devices on the Staff Baseline blueprint are personal equipment checked out to a
# person. Venue Baseline, Ministry Baseline, Apple TVs, MacOS Servers and Default Baseline
# are shared campus or ministry hardware that happens to carry a staff member's name as
# its administrative contact, which is not the same as being issued to them.
STAFF_BLUEPRINT = "Staff Baseline"
dev_by_user = collections.defaultdict(list)
skipped_blueprints = collections.Counter()
for d in devices:
    if not d["Device User"].strip(): continue
    if d["Blueprint Name"].strip() != STAFF_BLUEPRINT:
        skipped_blueprints[d["Blueprint Name"].strip() or "(blank)"] += 1
        continue
    dev_by_user[nkey(d["Device User"])].append(d)

def assets_for(disp):
    out = []
    if nkey(disp) in AMBIG_DISPLAY and dev_by_user.get(nkey(disp)):
        out.append(f"** {len(dev_by_user[nkey(disp)])} device(s) below are recorded in Kandji "
                   f"against the name '{disp}', which two AD accounts share - verify ownership **")
    for d in dev_by_user.get(nkey(disp), []):
        nm = d["Device Name"].strip() or d["Model Name"].strip() or "(unnamed)"
        sn = d["Device Serial Number"].strip() or "(no serial)"
        tag = d["Asset Tag"].strip()
        out.append(f"{nm} [{sn}]" + (f" (tag {tag})" if tag else ""))
    return sorted(out)

# ---------------- review CSV ----------------
with open(os.path.join(OUT, "AD-Name-Mapping-REVIEW.csv"), "w", newline="", encoding="utf-8-sig") as fh:
    w = csv.writer(fh)
    w.writerow(["Approved", "RosterRow", "RosterName", "Campus", "TargetOU", "Title", "Tier",
                "Confidence", "MatchVia", "MatchedDisplayName", "SamAccountName",
                "MatchedUPN", "AltCandidates", "Notes"])
    for p in roster:
        m = p["match"]
        alt = "; ".join(f"{c['a']['displayName']} <{c['a']['_local']}> {c['score']}/{c['via']}"
                        for c in p["cands"][1:4])
        n = []
        if p["dupe_of"]: n.append(f"DUPLICATE of roster row {p['dupe_of']}")
        if p["ou_note"]: n.append(p["ou_note"])
        if p["tier"] == "NO_STAFF_ACCOUNT":
            n.append("no staff account matched - create one, or confirm this role has no login")
        if p["tier"] == "AMBIGUOUS":
            n.append("INITIAL COLLISION or shared display name - confirm which account, may need first.last")
        if p["tier"] == "REVIEW":
            n.append("VERIFY SAME PERSON before approving")
        if m and m["why"]: n.append(m["why"])
        if p["mgr_note"]: n.append(f"{p['mgr_note']}: '{p['reports_to']}'")
        if p["pin_note"]: n.append("PINNED: " + p["pin_note"])
        if p["role_rank"] > 0:
            n.append("SECONDARY role - AD keeps the primary row for this account, this row is not written")
        elif m and any(q["role_rank"] > 0 for q in claimed.get(m["a"]["_local"], [])):
            n.append("PRIMARY role of a multi-role person")
        if nkey(p["reports_to"]) == nkey(f"{p['first']} {p['last']}"):
            n.append("roster lists this person as their own manager")
        if p["title"] in ("??", "?", "-", "TBD") or not p["title"]:
            n.append(f"title is placeholder text: '{p['title']}'")
        w.writerow(["Y" if p["tier"] == "AUTO" else "", p["row"], p["raw"], p["campus"], p["ou"],
                    p["title"], p["tier"], p["conf"], m["via"] if m else "",
                    m["a"]["displayName"] if m else "", m["a"]["_local"] if m else "",
                    m["a"]["userPrincipalName"] if m else "", alt, " | ".join(n)])

# ---------------- PowerShell input CSV ----------------
with open(os.path.join(OUT, "AD-Updates.csv"), "w", newline="", encoding="utf-8-sig") as fh:
    w = csv.writer(fh)
    w.writerow(["SamAccountName", "UserPrincipalName", "RosterName", "ADDisplayName", "Title",
                "Department", "Campus", "CompanyName", "OfficeName", "TargetOU", "EmployeeType",
                "ManagerSam", "ReportsToName", "Tier", "Confidence", "RosterRow"])
    for p in roster:
        if p["tier"] != "AUTO" or p["role_rank"] != 0: continue
        a = p["match"]["a"]
        w.writerow([a["_local"], a["userPrincipalName"], p["raw"], a["displayName"], p["title"],
                    p["dept"], p["campus"], CAMPUS_COMPANY.get(p["campus"], ""),
                    CAMPUS_DISPLAY.get(p["campus"], p["campus"]),
                    p["ou"], p["ftpt"], p["mgr_sam"], p["reports_to"],
                    p["tier"], p["conf"], p["row"]])

# ---------------- name corrections from AD ----------------
# AD is treated as authoritative for SPELLING. It is not authoritative for which name a
# person goes by: displayName routinely carries a preferred form (Chris for Christopher,
# Cookie for Eldora), and overwriting the roster with that would destroy the legal name.
# So spelling gets corrected, preferred names get their own column, and anything that is
# neither is held for a human.
#
# The account name is a better witness than displayName. A roster surname that looks like a
# typo against displayName may be corroborated by the samAccountName - in which case
# displayName is what is wrong, not the roster. Every surname correction must therefore be corroborated by
# the samAccountName before it is applied.
NICKNAME_PAIRS = set()
for _k, _vs in NICK.items():
    for _v in _vs:
        NICKNAME_PAIRS.add((_k, _v)); NICKNAME_PAIRS.add((_v, _k))

def classify_name(p):
    """-> (action, category, new_first, new_last, detail)"""
    m = p["match"]
    if not m: return None
    ad = m["a"]["_toks"]
    if not ad: return None
    ad_first, ad_last = " ".join(ad[:-1]), ad[-1]
    sam = m["a"]["_local"]
    rf, rl = alpha(p["first"]), alpha(p["last"])
    af, al = alpha(ad_first), alpha(ad_last)
    nf, nl, cats, det = p["first"], p["last"], [], []

    # ---- surname
    if rl and al and rl != al:
        parts = [alpha(x) for x in re.split(r"[-\s]+", p["last"]) if alpha(x)]
        sim = difflib.SequenceMatcher(None, rl, al).ratio()
        if al in parts and len(parts) > 1:
            cats.append("SURNAME TRUNCATED IN AD"); det.append(
                f"AD carries only '{ad_last}' of '{p['last']}' - the roster is more complete, "
                f"so AD is the side to fix")
        elif sim >= 0.80:
            # Corroborate against the account name, reconstructing it from each candidate
            # spelling. Substring tests are not safe here: a shorter surname can sit inside
            # a longer one and would wrongly vouch for it.
            def builds(given, surname):
                g = alpha(given)
                if not g or not surname: return set()
                return {g[0] + surname, given + "." + surname if False else g + "." + surname,
                        g + surname[0], surname, g}
            ad_builds = builds(af, al) | builds(alpha(p["first"]), al)
            rs_builds = builds(af, rl) | builds(alpha(p["first"]), rl)
            ad_ok, rs_ok = sam in ad_builds, sam in rs_builds
            if ad_ok and not rs_ok:
                nl = ad_last.capitalize() if ad_last.islower() else ad_last
                cats.append("SURNAME CORRECTED"); det.append(
                    f"'{p['last']}' -> '{nl}', confirmed by account '{sam}'")
            elif rs_ok and not ad_ok:
                cats.append("AD DISPLAYNAME WRONG"); det.append(
                    f"AD shows '{ad_last}' but the account is '{sam}', which reconstructs from "
                    f"the roster spelling '{p['last']}' - correct AD, not the roster")
            else:
                cats.append("SURNAME UNVERIFIABLE"); det.append(
                    f"roster '{p['last']}' vs AD '{ad_last}'; account '{sam}' corroborates neither")
        else:
            cats.append("SURNAME DIFFERENT"); det.append(
                f"roster '{p['last']}' vs AD '{ad_last}' - name change or wrong match, needs a person")

    # ---- given name
    if rf and af and rf != af:
        sim = difflib.SequenceMatcher(None, rf, af).ratio()
        is_nick = ((rf, af) in NICKNAME_PAIRS or af.startswith(rf) or rf.startswith(af)
                   or len(af) < len(rf) - 1 or sim < 0.6)
        if is_nick:
            cats.append("PREFERRED NAME"); det.append(
                f"AD shows '{ad_first}', roster has '{p['first']}' - recorded as a preferred "
                f"name, legal name left alone")
        elif sim >= 0.80:
            nf = ad_first.title() if ad_first.islower() else ad_first
            cats.append("GIVEN NAME CORRECTED"); det.append(f"'{p['first']}' -> '{nf}'")
        else:
            cats.append("GIVEN NAME DIFFERENT"); det.append(
                f"roster '{p['first']}' vs AD '{ad_first}' - needs a person")
    if not cats: return None
    changed = (nf != p["first"] or nl != p["last"])
    return ("apply" if changed else "hold", "; ".join(cats), nf, nl, "; ".join(det))

for p in roster:
    p["name_fix"] = classify_name(p) if p["tier"] in ("AUTO", "REVIEW") else None
    p["corrected_name"] = p["raw"]
    p["preferred"] = ""
    p["override_note"] = ""
    if p["row"] in ROSTER_NAME_OVERRIDES:
        newname, why = ROSTER_NAME_OVERRIDES[p["row"]]
        p["corrected_name"] = newname
        p["override_note"] = why
        if p["match"]:
            p["preferred"] = (p["match"]["a"]["displayName"]
                              if nkey(p["match"]["a"]["displayName"]) != nkey(newname) else "")
        continue
    nfx = p["name_fix"]
    if nfx:
        action, cat, nf, nl, det = nfx
        if action == "apply":
            p["corrected_name"] = f"{nl}, {nf}" if "," in p["raw"] else f"{nf} {nl}"
            if p["alias"]: p["corrected_name"] += f" ({p['alias']})"
        if "PREFERRED NAME" in cat:
            p["preferred"] = p["match"]["a"]["displayName"]

with open(os.path.join(OUT, "AD-Name-Corrections.csv"), "w", newline="", encoding="utf-8-sig") as fh:
    w = csv.writer(fh)
    w.writerow(["Action", "Category", "RosterRow", "RosterName", "CorrectedName",
                "ADDisplayName", "SamAccountName", "PreferredName", "Detail"])
    for p in roster:
        if p["override_note"]:
            w.writerow(["APPLIED - confirmed in review", "ROSTER NAME OVERRIDE", p["row"], p["raw"],
                        p["corrected_name"],
                        p["match"]["a"]["displayName"] if p["match"] else "",
                        p["match"]["a"]["_local"] if p["match"] else "",
                        p["preferred"], p["override_note"]])
            continue
        if not p["name_fix"]: continue
        action, cat, nf, nl, det = p["name_fix"]
        if action == "apply" and p["tier"] != "AUTO":
            action_label = "APPLIED - but confirm the match first"
        elif action == "apply":
            action_label = "APPLIED"
        else:
            action_label = "HELD - needs a decision"
        w.writerow([action_label, cat, p["row"],
                    p["raw"], p["corrected_name"] if action == "apply" else "",
                    p["match"]["a"]["displayName"], p["match"]["a"]["_local"],
                    p["preferred"], det])

# ---------------- exceptions report ----------------
with open(os.path.join(OUT, "AD-Exceptions.csv"), "w", newline="", encoding="utf-8-sig") as fh:
    w = csv.writer(fh)
    w.writerow(["Category", "RosterRow", "RosterName", "SamAccountName", "Campus",
                "Department", "Title", "FT/PT", "ReportsTo", "Detail"])
    def emit(cat, p, detail):
        m = p["match"]
        w.writerow([cat, p["row"], p["raw"], m["a"]["_local"] if m else "", p["campus"],
                    p["dept"], p["title"], p["ftpt"], p["reports_to"], detail])
    for loc, ps in sorted(dual):
        for q in ps:
            emit("DUAL ROLE " + ("(primary - written to AD)" if q["role_rank"] == 0
                                 else "(secondary - NOT written)"), q,
                 f"{len(ps)} roster rows share {loc}: " +
                 "; ".join(f"r{z['row']} {z['dept']}/{z['title']} {z['ftpt']}" for z in ps))
    for p in roster:
        if p["tier"] == "NO_STAFF_ACCOUNT":
            emit("NO STAFF ACCOUNT", p, "no account matched any naming pattern, including "
                                        "surname-spelling variants - create a staff account, "
                                        "or confirm this role does not get a login")
        elif p["tier"] == "AMBIGUOUS":
            emit("AMBIGUOUS MATCH", p,
                 "candidates: " + "; ".join(f"{c['a']['displayName']} <{c['a']['_local']}> {c['score']}"
                                            for c in p["cands"][:4]))
        elif p["tier"] == "REVIEW":
            emit("LOW CONFIDENCE", p,
                 f"best guess {p['match']['a']['displayName']} <{p['match']['a']['_local']}> "
                 f"scored only {p['conf']} - initials collide, verify before approving")
        if p["mgr_note"]:
            cat = "MANAGER NEAR-MISS" if "near-miss" in p["mgr_note"] else "MANAGER UNRESOLVED"
            impact = "" if p["tier"] == "AUTO" else " (this person has no AD account either, so no impact)"
            emit(cat, p, f"'{p['reports_to']}': {p['mgr_note']}{impact}")
        if nkey(p["reports_to"]) == nkey(f"{p['first']} {p['last']}"):
            emit("SELF-MANAGER", p, "roster lists this person as their own manager")
        if not p["title"] or p["title"] in ("??", "?", "-", "TBD"):
            emit("PLACEHOLDER TITLE", p, f"title cell contains '{p['title']}'")
        if p["campus"] and not p["ou"]:
            emit("UNMAPPED CAMPUS", p, f"campus '{p['campus']}' has no OU mapping")
        if p["pin_note"]:
            emit("PINNED IDENTITY", p, p["pin_note"])
        if p["tier"] == "AUTO" and nkey(p["match"]["a"]["displayName"]) in AMBIG_DISPLAY:
            emit("SHARED DISPLAY NAME", p,
                 f"'{p['match']['a']['displayName']}' is used by 2 internal accounts; "
                 f"Kandji device attribution for this name is unverifiable")

# ---------------- stale / offboarding report ----------------
roster_sams = {p["match"]["a"]["_local"] for p in roster if p["tier"] == "AUTO"}
with open(os.path.join(OUT, "AD-Stale-Accounts.csv"), "w", newline="", encoding="utf-8-sig") as fh:
    w = csv.writer(fh)
    w.writerow(["SamAccountName", "UserPrincipalName", "DisplayName", "CompanyName",
                "LastSignIn", "DaysSinceSignIn", "OnRoster", "Recommendation"])
    stale = [a for a in pool if (a["_age"] is None or a["_age"] >= STALE_DAYS)]
    for a in sorted(stale, key=lambda x: -(x["_age"] or 99999)):
        on = a["_local"] in roster_sams
        if on:
            rec = "ON ROSTER but dormant - confirm still employed"
        elif a["_age"] is None:
            rec = "never signed in - confirm the account is needed"
        else:
            rec = f"no sign-in in {a['_age']} days and not on the roster - OFFBOARD, do not flag Volunteer"
        w.writerow([a["_local"], a["userPrincipalName"], a["displayName"], a["companyName"],
                    a["_last"] or "never", a["_age"] if a["_age"] is not None else "",
                    "Y" if on else "N", rec])

# ---------------- enriched workbook ----------------
ws.insert_cols(7, 6)
for col, hdr in zip("GHIJKL", ["Preferred Name", "Staff Account", "IT Assets",
                               "Passwordless Migration Status", "AD Account",
                               "AD Match Confidence"]):
    ws[f"{col}1"] = hdr

# spelling corrections taken from AD, written straight into the Name column
BLUE = PatternFill("solid", fgColor="DDEBF7")
name_fixes = 0
for p in roster:
    if p["corrected_name"] != p["raw"]:
        c = ws.cell(p["row"], 1)
        c.value = p["corrected_name"]; c.fill = BLUE
        name_fixes += 1
    if p["preferred"]:
        ws.cell(p["row"], 7, p["preferred"])
AMBER = PatternFill("solid", fgColor="FFF2CC")
RED = PatternFill("solid", fgColor="FCE4E4")
GREEN = PatternFill("solid", fgColor="E2EFDA")
for p in roster:
    r = p["row"]; m = p["match"]
    if p["tier"] == "AUTO":
        disp, upn = m["a"]["displayName"], m["a"]["userPrincipalName"].lower()
        a = assets_for(disp)
        ws.cell(r, 8, "Yes")
        ws.cell(r, 8).fill = GREEN
        ws.cell(r, 9, "\n".join(a) if a else "None")
        ws.cell(r, 10, "Pending migration" if upn in pw_upn else "Migrated")
        ws.cell(r, 11, upn)
        ws.cell(r, 12, f"{p['conf']} ({m['via']})")
    elif p["tier"] == "NO_STAFF_ACCOUNT":
        for c, v in ((8, "No"), (9, "No staff account"), (10, "No staff account"),
                     (11, ""), (12, "0")):
            cell = ws.cell(r, c); cell.value = v; cell.fill = RED
    else:
        lbl = "Needs review" if p["tier"] == "REVIEW" else "Ambiguous"
        guess = m["a"]["userPrincipalName"].lower() if m else ""
        for c, v in ((8, f"Unconfirmed - {lbl}"),
                     (9, f"Unconfirmed - {lbl}"),
                     (10, f"Unconfirmed - {lbl}"),
                     (11, (f"candidate: {guess}" if guess else lbl)),
                     (12, f"{p['conf']}" + (f" ({m['via']})" if m else ""))):
            cell = ws.cell(r, c); cell.value = v; cell.fill = AMBER

# ---------------- merge volunteers into the main sheet ----------------
# Everyone with an internal @example.org account who is not on the staff roster,
# appended to the same sheet with their standing recorded in the FT/PT column. Three values
# rather than one: an account dormant for two years and a shared mailbox are both "not
# staff", but they are not the same thing as someone who moved into volunteering, and
# collapsing them would hide that.
# Matched against displayName AND the account name: service accounts often have perfectly
# human-looking display names, so a display-name-only test calls them volunteers.
SHARED_RX = re.compile(r"\b(team|mailbox|helpdesk|rooms?|calendar|workflow|kiosk|shared|scan|"
                       r"printer|signage|display|referrals|onboarding|maintenance|support|"
                       r"admin|automation|robots|test|info|security|bot|scheduler|alerts?|"
                       r"notifications?|noreply|donotreply|service|daemon|sync|monitor)\b", re.I)
SHARED_SAM_RX = re.compile(r"(bot|scheduler|alerts?|notif|noreply|donotreply|svc|service|"
                           r"daemon|sync|monitor|^hve-|^sm_|^adm-|^\$)", re.I)

# AD companyName back to the roster's campus shorthand, so the CAMPUS column stays uniform.
CO_TO_CAMPUS = [
    # AD Company value -> roster site code. Most specific first; the first substring hit wins.
    ("central", "Central"), ("north", "North"), ("south", "South"), ("hq", "HQ"),
]
def campus_from_company(co):
    c = (co or "").lower()
    for frag, camp in CO_TO_CAMPUS:
        if frag in c: return camp
    return ""

roster_sams_final = {p["match"]["a"]["_local"] for p in roster if p["tier"] == "AUTO"}
vol = []
for a in pool:
    if a["_local"] in roster_sams_final: continue
    devs = [d for d in assets_for(a["displayName"]) if not d.startswith("**")]
    shared = (bool(SHARED_RX.search(a["displayName"]))
               or bool(SHARED_SAM_RX.search(a["_local"]))
               or len(a["_toks"]) < 2)
    dormant = a["_age"] is None or a["_age"] >= STALE_DAYS
    disabled = a["_local"] in disabled_sams
    age_txt = "no interactive sign-in ever" if a["_age"] is None else f"no interactive sign-in in {a['_age']} days"

    # Disabled outranks dormant: a disabled account is a decision someone already made, and
    # calling it dormant describes the symptom instead of the state.
    if disabled and shared:
        status, note = "Shared Mailbox (disabled)", "not a person, and already disabled - decommission or delete"
    elif disabled:
        extra = disabled_detail.get(a["_local"], "")
        status = "Disabled - pending offboarding"
        note = f"account is disabled; {age_txt}" + (" | " + extra if extra else "")
    elif shared:
        status, note = "Shared Mailbox", "not a person - the sync never labels these"
    elif dormant:
        status, note = "Volunteer (dormant)", f"{age_txt} - check whether it should be offboarded"
    else:
        status, note = "Volunteer", "active account, not on the staff roster"
    if a["_noninteractive_only"]:
        note += (f" | still shows non-interactive activity to {a['_last_noninteractive']}"
                 f" (token refresh, not a person signing in)")
    if not disabled_known:
        note += " | enabled state unknown - AD-Disabled-Accounts.csv was not supplied"
    vol.append({"a": a, "devs": devs, "status": status, "note": note})

order = {"Volunteer": 0, "Volunteer (dormant)": 1, "Disabled - pending offboarding": 2,
         "Shared Mailbox": 3, "Shared Mailbox (disabled)": 4}
vol.sort(key=lambda v: (order[v["status"]], -len(v["devs"]), v["a"]["displayName"]))

AMBER_V = PatternFill("solid", fgColor="FFF2CC")
FLAG_V  = PatternFill("solid", fgColor="FFC7CE")
GREY_V  = PatternFill("solid", fgColor="EDEDED")
row = ws.max_row + 1
flagged = 0
for v in vol:
    a, devs = v["a"], v["devs"]
    t = a["_toks"]
    name = f"{t[-1].title()}, {' '.join(x.title() for x in t[:-1])}" if len(t) >= 2 else a["displayName"]
    holds = bool(devs) and v["status"] != "Shared Mailbox"
    note = v["note"]
    if holds and v["status"].startswith("Disabled"):
        holds = True   # a disabled account still holding equipment is the sharper version
    if holds:
        flagged += 1
        note = "HOLDS STAFF EQUIPMENT while off the staff roster - confirm it should stay checked out. " + note
    if a.get("onPremisesSyncEnabled") != "True":
        note += (" | CLOUD-ONLY account, not synced from AD - the sync script cannot reach it; "
                 "any change must be made in Entra")
    vals = [name, campus_from_company(a["companyName"]), "", "", "", v["status"],
            # only a genuinely different form counts as a preferred name; "Smith, Jane"
            # against "Jane Smith" is the same name reordered
            a["displayName"] if sorted(toks(name)) != sorted(a["_toks"]) else "",
            "Yes", "\n".join(devs) if devs else "None",
            "Pending migration" if a["userPrincipalName"].lower() in pw_upn else "Migrated",
            a["userPrincipalName"].lower(), note]
    for i, val in enumerate(vals, start=1):
        c = ws.cell(row, i); c.value = val
        c.fill = FLAG_V if holds else (GREY_V if v["status"] == "Shared Mailbox" else AMBER_V)
    row += 1

# A roster member whose account is disabled keeps their FT/PT - the roster says they are
# staff. The contradiction goes in the notes column, where it can be acted on.
staff_disabled = 0
for p in roster:
    if p["tier"] != "AUTO": continue
    if p["match"]["a"]["_local"] not in disabled_sams: continue
    staff_disabled += 1
    c = ws.cell(p["row"], 12)
    extra = disabled_detail.get(p["match"]["a"]["_local"], "")
    c.value = ("DISABLED but on the staff roster - resolve whether the account was disabled "
               "in error or the roster is stale" + (" | " + extra if extra else ""))
    c.fill = FLAG_V

ws.cell(1, 12, "Volunteer / Account Notes")
print("volunteer rows merged:", len(vol), "|", dict(collections.Counter(v["status"] for v in vol)))
print("off roster but holding Staff Baseline equipment:", flagged)
_cloud = [v for v in vol if v["a"].get("onPremisesSyncEnabled") != "True"]
print("cloud-only non-roster accounts (unreachable by the sync):", len(_cloud),
      dict(collections.Counter(v["status"] for v in _cloud)))
print("roster staff whose account is disabled:", staff_disabled)
print("dormant only by interactive signal:", sum(1 for a in pool if a["_noninteractive_only"]))
print("sheet1 rows now:", ws.max_row)

hf = PatternFill("solid", fgColor="1F3864")
for col in range(1, 13):
    c = ws.cell(1, col); c.font = Font(bold=True, color="FFFFFF"); c.fill = hf
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
for col, wd in (("G", 20), ("H", 15), ("I", 48), ("J", 24), ("K", 34), ("L", 20)):
    ws.column_dimensions[col].width = wd
for r in range(2, ws.max_row + 1):
    for col in range(7, 13):
        ws.cell(r, col).alignment = Alignment(vertical="top", wrap_text=True)
ws.auto_filter.ref = f"A1:L{ws.max_row}"
ws.freeze_panes = "A2"
wb.save(os.path.join(OUT, "Staff-Roster-ENRICHED.xlsx"))

# ---------------- summary ----------------
t = collections.Counter(p["tier"] for p in roster)
auto = [p for p in roster if p["tier"] == "AUTO"]
print("roster rows:", len(roster))
print("tiers:", dict(t))
print("via:", dict(collections.Counter(p["match"]["via"] for p in auto)))
print("assets found (Staff Baseline only):", sum(1 for p in auto if assets_for(p["match"]["a"]["displayName"])))
print("devices excluded by blueprint:", dict(skipped_blueprints))
print("name spellings corrected from AD:", name_fixes)
print("preferred names recorded:", sum(1 for p in roster if p["preferred"]))
print("pending passwordless:", sum(1 for p in auto if p["match"]["a"]["userPrincipalName"].lower() in pw_upn))
print("duplicate roster rows:", sum(1 for p in roster if p["dupe_of"]))
print("campus remapped:", sum(1 for p in roster if p["campus"] in REMAPPED))
print("internal AD members:", len(pool), "| unclaimed by roster:", len(pool) - len({p["match"]["a"]["_local"] for p in auto}))
print("dual-role people:", len(dual), "| secondary rows suppressed:", sum(len(ps) - 1 for _, ps in dual))
print("pinned / reviewed identities:", sum(1 for p in roster if p["pin_note"]))
print("roster name overrides:", sum(1 for p in roster if p["override_note"]))
print("forced secondary rows:", sorted(SECONDARY_ROLE_ROWS))
print("settled as no-account:", sorted(NO_ACCOUNT_ROWS))
print("shared display names:", len(AMBIG_DISPLAY), "->", sorted(AMBIG_DISPLAY))
_st = [a for a in pool if a["_age"] is None or a["_age"] >= STALE_DAYS]
print(f"stale accounts (>={STALE_DAYS}d or never):", len(_st),
      "| of those not on roster:", sum(1 for a in _st if a["_local"] not in roster_sams))
for tier in ("AMBIGUOUS", "REVIEW"):
    print(f"\n-- {tier} --")
    for p in roster:
        if p["tier"] == tier:
            print(f"  {p['raw']:32s} -> {p['match']['a']['displayName']:24s} "
                  f"<{p['match']['a']['_local']}> {p['conf']} {p['match']['via']}")
